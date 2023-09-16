from PIL import Image
from openpyxl import Workbook, load_workbook
import openpyxl.worksheet.worksheet
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

workbook_ref = openpyxl.workbook.Workbook
worksheet_ref = openpyxl.worksheet.worksheet.Worksheet

def Rgb_to_Hex(r: int, g: int, b: int) -> str :
    """ Converts RGB Color to Hex

    Args:
        r : Red
        g : Green
        b : Blue
    Returns:
        Hex Value
    """
    return '{:02X}{:02X}{:02X}'.format(r, g, b)

def ColorCodes_to_Excel(img_name: str , excel_name: str) -> None:
    """ Put Hex Color value in a Cell(Excel) corresponding to Pixel Position(Image)

     Args:
         img_name : Image name in the Browser
         excel_name : Name of Excel Workbook

    Returns:
        None
    """
    wb: workbook_ref  = Workbook()
    ws: worksheet_ref = wb.active

    im = Image.open(img_name, "r")
    X, Y = im.size # Get Image Dimensions
    c = im.convert("RGB")

    for y in range(Y):
        row=[]
        for x in range(X):
            color = c.getpixel((x,y)) # Get RGB Color of each pixel
            hexValue = Rgb_to_Hex(*color)
            row.append(hexValue) # Make a list of Hex Values of pixels in one row (Image)
        ws.append(row) # Adding the Hex List to each row in Excel Sheet
    wb.save(excel_name)
    wb.close()

def ColoringCells_in_Excel(excel_name: str) -> None:
    """ Color each cell with its corresponding value (Hex)

    Args:
        excel_name: Name of Excel Workbook

    Returns:
        None
    """
    wb: workbook_ref = load_workbook(excel_name)
    ws: worksheet_ref = wb.active

    for Cells in ws.iter_cols(): # Get Cells in Columns
        for c in Cells: # Get each Cell in a Column
            color = PatternFill(fill_type='solid', start_color=c.value) # Get color from Cell value (Hex)
            cell: Cell = c
            cell.fill = color # Fill Color Cell

    Format_Width_Height(ws)

    wb.save(excel_name)
    wb.close()

def get_ColumnNames(ws: worksheet_ref) -> list:
    """ Get the name of All used Columns

    Args:
        ws: Excel WorkSheet

    Returns:
        List of Columns Names
    """
    Columns = []
    for Cells in ws.iter_cols(): # Get Cells in Columns
        for c in Cells:
            s: str = c.coordinate # Get Cell position
            s = ''.join((z for z in s if not z.isdigit())) # Remove any digits (Row No) in cell position
            Columns.append(s) # Add all Column names
    Columns = list(set(Columns)) # Remove All Duplicates in the Columns list
    return Columns

def Format_Width_Height(ws: worksheet_ref) -> None:
    """ Format Columns_Width and Row_Height in all Cells

    Args:
        ws: Excel WorkSheet

    Returns:
        None
    """
    for col in get_ColumnNames(ws=ws): # Get each name in columns
        ws.column_dimensions[col].width = 3 # Format Column_Width

    for i in range(ws.max_row+1): # Get each number in rows
        ws.row_dimensions[i].height = 15 # Format Row_Height

def main():
    imgName = "Insta.jpg"
    excelName = "Instagram.xlsx"
    ColorCodes_to_Excel(imgName, excelName)
    ColoringCells_in_Excel(excelName)

if __name__ == "__main__":
    main()
